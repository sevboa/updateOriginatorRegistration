import re
import sys

from openpyxl import load_workbook

sys.path.append('./class/originatorImporter')
from Importer import importer
from OriginatorMts import originatorMts

class importerBt(importer):
    
    def load(self, extension, providerCode, fileName):
        if extension == 'xlsx' and providerCode in ('BT'):
            wb = load_workbook(filename = './originators/' + fileName)
            for sheetName in wb.sheetnames[:1]:
                sheet = wb[sheetName]
                i = int()
                while True:
                    i+=1
                    if sheet['A' + str(i)].value == None:
                        break
                    try:
                        if sheet['A' + str(i)].value == None:
                            break
                        for stringObject in sheet['A' + str(i):'D' + str(i)]:
                                self.input(stringObject)
                    except AttributeError:
                        break
        else:
            print('Неподдерживаемое расширение входящего файла!')
            quit()

    def outerOriginatorsAppend(self, stringObject):
        if type(stringObject[2].value) == type(int()):
            originator = originatorMts(stringObject, self.RejectMessageMts)
            if self.GlobalOriginators.isdisjoint({originator.Originator + ';' + str(originator.ProviderId)}):
                self.Originators.add(originator)

    def input(self, stringObject):
        self.outerOriginatorsAppend(stringObject)
