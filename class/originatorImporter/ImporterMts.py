import re
import sys
import csv

from openpyxl import load_workbook

sys.path.append('./class/originatorImporter')
from Importer import importer
from OriginatorMts import originatorMts

class importerMts(importer):
    
    Statuses = list()
    StatusesCache = dict()
    
    def loadConfig(self):
        self.Statuses.extend(self.Config['mts']['statuses'])
        self.createStatusesCache()

    def load(self, extension, providerCode, fileName):
        if extension == 'xlsx' and providerCode in ('MTS'):
            wb = load_workbook(filename = './originators/' + fileName)
            for sheetName in wb.sheetnames[:1]:
                sheet = wb[sheetName]
                i = int()
                while True:
                    i+=1
                    if sheet['A' + str(i)].value == None:
                        break
                    try:
                        if sheet['A' + str(i)].value == None:
                            break
                        for stringObject in sheet['A' + str(i):'D' + str(i)]:
                            if re.sub(r'[^\d]+', '', str(stringObject[2].value)) != '':
                                if 'MTS_error_' in fileName:
                                    stateMessage = str(stringObject[3].value)
                                    statusAndIsIncorrectInn = self.StatusesCache.get(stateMessage)
                                    if statusAndIsIncorrectInn == None:
                                        print(str(stringObject[3].value))
                                        print('Неожиданный статус!')
                                        #writeCsv.writerow([stringObject[3].value, '', ''])
                                        self.StatusesCache.update({stringObject[3].value : '0' + ';' + '0'})
                                    elif statusAndIsIncorrectInn != '0;0':
                                        self.input(stringObject, True)
                                else:
                                    self.input(stringObject)
                    except AttributeError:
                        break
        else:
            print('Неподдерживаемое расширение входящего файла!')
            quit()
    
    def input(self, stringObject, rejected=False):
        if rejected == True:
            originator = originatorMts(stringObject, self.StatusesCache)
        else:
            originator = originatorMts(stringObject)
        if self.GlobalOriginatorsCache.isdisjoint({originator.Originator + ';' + str(originator.ProviderId)}):
            self.Originators.add(originator)

    def createStatusesCache(self):
        for status in self.Statuses:
            self.StatusesCache.update({status['text'] : str(status['registration_status_id']) + ';' + str(status['is_incorrect_inn'])})
        

