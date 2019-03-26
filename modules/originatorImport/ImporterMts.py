import re
import csv

from openpyxl import load_workbook

from modules.originatorImport.Importer import importer
from modules.originatorImport.OriginatorMts import originatorMts

class importerMts(importer):
    
    def loadConfig(self):
        self.createMtsStatusesCache()
            
    def load(self, extension, providerCode, fileName):
        
        if extension == 'xlsx' and providerCode in ('MTS'):
            wb = load_workbook(filename = './originators/' + fileName)
            for sheetName in wb.sheetnames[:1]:
                sheet = wb[sheetName]
                i = int()
                while True:
                    i+=1
                    if sheet['A' + str(i)].value == None:
                        break
                    if sheet['A' + str(i)].value == None:
                        break
                    for stringObject in sheet['A' + str(i):'D' + str(i)]:
                        if re.sub(r'[^\d]+', '', str(stringObject[2].value)) != '':
                            if 'MTS_error_' in fileName:
                                originator = originatorMts(stringObject)
                            else:
                                originator = originatorMts(stringObject, False)
                            self.Originators.add(originator)
        else:
            print('Неподдерживаемое расширение входящего файла!')
            quit()

    def createMtsStatusesCache(self):
        self.originatorConfig['mts']['StatusesCache'] = dict()
        for status in self.originatorConfig['mts']['Statuses']:
            self.originatorConfig['mts']['StatusesCache'].update({status['text'] : str(status['registration_status_id']) + ';' + str(status['is_incorrect_inn'])})

    
        

